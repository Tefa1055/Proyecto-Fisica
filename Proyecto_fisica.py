import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import linregress
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas as pdf_canvas
import tempfile
import os

ventana = tk.Tk()
ventana.title("Análisis de Movimiento Armónico Simple")
ventana.geometry("900x700")
ventana.configure(bg="#f0f0f0")

canvas_scroll = tk.Canvas(ventana, bg="#f0f0f0")
scrollbar = tk.Scrollbar(ventana, orient="vertical", command=canvas_scroll.yview)
scrollable_frame = tk.Frame(canvas_scroll, bg="#f0f0f0")
scrollable_frame.bind("<Configure>", lambda e: canvas_scroll.configure(scrollregion=canvas_scroll.bbox("all")))
canvas_scroll.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas_scroll.configure(yscrollcommand=scrollbar.set)
canvas_scroll.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

datos = {}
canvas = None

def cargar_excel():
    global datos
    archivo = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
    if not archivo:
        return
    try:
        from openpyxl import load_workbook
        libro = load_workbook(filename=archivo, data_only=True)
        hoja = libro.active
        t, x, v, Ec, Ep = [], [], [], [], []
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if all(c is not None for c in fila[:5]):
                t.append(float(fila[0]))
                x.append(float(fila[1]))
                v.append(float(fila[2]))
                Ec.append(float(fila[3]))
                Ep.append(float(fila[4]))
        datos = {
            "t": np.array(t),
            "x": np.array(x),
            "v": np.array(v),
            "Ec": np.array(Ec),
            "Ep": np.array(Ep),
            "Etot": np.array(Ec) + np.array(Ep)
        }
        messagebox.showinfo("Éxito", "Datos cargados correctamente.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo leer el archivo:\n{e}")

def analizar_grafica(opcion):
    if not datos:
        return "❌ No hay datos cargados."

    t = datos["t"]
    explicacion = ""

    if opcion == "Posición vs Tiempo":
        x = datos["x"]
        amplitud = (np.max(x) - np.min(x)) / 2
        media = np.mean(x)
        cruces = np.sum(np.diff(np.sign(x - media)) != 0)
        desvio = np.std(x)

        explicacion += f"🔹 Amplitud estimada: {amplitud:.2f} m\n"
        explicacion += f"🔹 Valor medio: {media:.2f} m\n"
        explicacion += f"🔹 Cruces por el eje: {cruces}\n"

        indices_cruces = np.where(np.diff(np.sign(x - media)))[0]
        if len(indices_cruces) > 1:
            tiempos_cruces = t[indices_cruces]
            difs = np.diff(tiempos_cruces)
            positivos = difs[difs > 0]
            if len(positivos) > 0:
                periodo = np.mean(positivos) * 2
                frecuencia = 1 / periodo
                explicacion += f"🔹 Período estimado: {periodo:.2f} s | Frecuencia: {frecuencia:.2f} Hz\n"

        if cruces >= 4 and desvio > 0.1 and amplitud > 0.1:
            explicacion += "✅ Posición: El movimiento oscila regularmente alrededor de un punto medio, con amplitud constante y múltiples cruces. Es MAS."
        else:
            explicacion += "⚠️ Posición: No hay suficientes cruces o la amplitud es baja. No parece MAS."

    elif opcion == "Velocidad vs Tiempo":
        v = datos["v"]
        desvio = np.std(v)
        v_max = np.max(v)
        v_min = np.min(v)
        cruces = np.sum(np.diff(np.sign(v - np.mean(v))) != 0)

        explicacion += f"🔹 Velocidad máxima: {v_max:.2f} m/s\n"
        explicacion += f"🔹 Velocidad mínima: {v_min:.2f} m/s\n"
        explicacion += f"🔹 Cruces por el eje: {cruces}\n"

        if cruces >= 4 and desvio > 0.1:
            explicacion += "✅ Velocidad: comportamiento periódico típico de MAS."
        else:
            explicacion += "⚠️ Velocidad: no se observan suficientes cruces ni variación."

    elif opcion == "Energías (Ec, Ep, Etot) vs Tiempo":
        Ec = datos["Ec"]
        Ep = datos["Ep"]
        Etot = datos["Etot"]
        energia_constante = np.allclose(Etot, np.mean(Etot), atol=0.2)
        explicacion += f"🔹 Ec promedio: {np.mean(Ec):.2f} J\n"
        explicacion += f"🔹 Ep promedio: {np.mean(Ep):.2f} J\n"
        explicacion += f"🔹 Etot promedio: {np.mean(Etot):.2f} J\n"
        if energia_constante:
            explicacion += "✅ Energía total constante → MAS ideal."
        else:
            explicacion += "⚠️ Energía total variable → sistema no ideal."

    elif opcion == "Regresión lineal de Energía Total":
        Etot = datos["Etot"]
        slope, intercept, r, _, _ = linregress(t, Etot)
        explicacion += f"🔹 Pendiente: {slope:.2e}\n"
        explicacion += f"🔹 Intersección: {intercept:.2f}\n"
        explicacion += f"🔹 R²: {r**2:.4f}\n"
        if abs(slope) < 1e-3:
            explicacion += "✅ Etot constante → conserva energía, posible MAS."
        else:
            explicacion += "⚠️ Etot cambia → no es MAS ideal."

    return explicacion

def mostrar_grafico():
    global canvas
    if not datos:
        messagebox.showwarning("Advertencia", "Primero carga un archivo Excel.")
        return
    opcion = combo.get()
    t = datos["t"]
    if canvas:
        canvas.get_tk_widget().destroy()
    fig = plt.Figure(figsize=(6, 4), dpi=100)
    ax = fig.add_subplot(111)
    if opcion == "Posición vs Tiempo":
        ax.plot(t, datos["x"], label="x(t)", color='blue')
        ax.set_ylabel("Posición (m)")
    elif opcion == "Velocidad vs Tiempo":
        ax.plot(t, datos["v"], label="v(t)", color='orange')
        ax.set_ylabel("Velocidad (m/s)")
    elif opcion == "Energías (Ec, Ep, Etot) vs Tiempo":
        ax.plot(t, datos["Ec"], label="Ec", color='green')
        ax.plot(t, datos["Ep"], label="Ep", color='blue')
        ax.plot(t, datos["Etot"], label="Etot", color='red')
        ax.set_ylabel("Energía (J)")
    elif opcion == "Regresión lineal de Energía Total":
        Etot = datos["Etot"]
        slope, intercept, _, _, _ = linregress(t, Etot)
        ax.scatter(t, Etot, label="Etot", color='red')
        ax.plot(t, slope * t + intercept, label=f"y = {slope:.2e}x + {intercept:.2f}", color='black')
        ax.set_ylabel("Energía Total (J)")
    ax.set_title(opcion)
    ax.set_xlabel("Tiempo (s)")
    ax.grid(True)
    ax.legend()
    canvas = FigureCanvasTkAgg(fig, master=frame_grafico)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)
    label_explicacion.config(text="🔎 Análisis:\n" + analizar_grafica(opcion))

def exportar_pdf():
    if not datos or not canvas:
        messagebox.showwarning("Error", "Primero muestra una gráfica.")
        return
    ruta_pdf = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("Archivo PDF", "*.pdf")])
    if not ruta_pdf:
        return
    temp_img = os.path.join(tempfile.gettempdir(), "grafica_temp.png")
    canvas.figure.savefig(temp_img)
    c = pdf_canvas.Canvas(ruta_pdf, pagesize=letter)
    width, height = letter
    c.setFont("Helvetica-Bold", 14)
    c.drawString(150, height - 40, "Informe de Análisis de MAS")
    c.drawImage(temp_img, 100, height - 300, width=400, height=200)
    y = height - 320
    for linea in label_explicacion.cget("text").split("\n"):
        if y < 50:
            c.showPage()
            y = height - 50
        c.drawString(40, y, linea)
        y -= 15
    c.save()
    os.remove(temp_img)
    messagebox.showinfo("PDF generado", f"Informe guardado correctamente:\n{ruta_pdf}")

# Interfaz gráfica
tk.Label(scrollable_frame, text="Análisis de Movimiento Armónico Simple", font=("Helvetica", 16), bg="#f0f0f0").pack(pady=10)
tk.Button(scrollable_frame, text="📁 Cargar archivo Excel", command=cargar_excel, font=("Helvetica", 12), bg="#007ACC", fg="white").pack(pady=5)
tk.Label(scrollable_frame, text="Selecciona la gráfica a mostrar:", font=("Helvetica", 12), bg="#f0f0f0").pack()
combo = ttk.Combobox(scrollable_frame, values=[
    "Posición vs Tiempo",
    "Velocidad vs Tiempo",
    "Energías (Ec, Ep, Etot) vs Tiempo",
    "Regresión lineal de Energía Total"
], state="readonly", width=50)
combo.current(0)
combo.pack(pady=5)
tk.Button(scrollable_frame, text="📊 Mostrar gráfica", command=mostrar_grafico, font=("Helvetica", 12), bg="#4CAF50", fg="white").pack(pady=10)
tk.Button(scrollable_frame, text="📄 Exportar análisis a PDF", command=exportar_pdf, font=("Helvetica", 12), bg="#3F51B5", fg="white").pack(pady=5)
frame_grafico = tk.Frame(scrollable_frame, bg="white", bd=2, relief="sunken")
frame_grafico.pack(padx=10, pady=10, fill="both", expand=True)
label_explicacion = tk.Label(scrollable_frame, text="", font=("Helvetica", 11), bg="#f0f0f0", justify="left", anchor="w", wraplength=850)
label_explicacion.pack(padx=20, pady=(0, 15), fill="x")

ventana.mainloop() 